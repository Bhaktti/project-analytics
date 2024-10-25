import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.chart import LineChart

# Load the dataset
file_name = './purchase_data.xlsx'
if os.path.exists(file_name):
    product_df = pd.read_excel(file_name, sheet_name='Sheet1')

    # Data Cleaning and Preparation
    product_df.drop_duplicates(keep='first', inplace=True, ignore_index=False)
    product_df['Gender'] = product_df['Gender'].replace('F', 'Female').replace('M', 'Male')
    product_df['Age brackets'] = product_df['Age'].apply(lambda x: 'Less than 30' if x <= 30 else ('Greater than 55' if x > 55 else '31 to 55'))
    product_df['Date'] = pd.to_datetime(product_df['Date'], errors='coerce')
    product_df['Date'] = product_df['Date'].dt.date
    product_df['Date brackets'] = product_df['Date'].apply(lambda date: '2023 and later' if date.year >= 2023 else ('2020 to 2022' if date.year >= 2020 else 'Before 2020'))
    product_df['Returns'].fillna(0, inplace=True)
    product_df['Returns'] = product_df['Returns'].astype(int)

    # Save cleaned data to a new sheet
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
        product_df.to_excel(writer, sheet_name='Working_Sheet', index=False)

    # Analysis
    churn_by_age = product_df.groupby('Age brackets').agg(
        Churned=('Churn', lambda x: (x == 1).sum()),
        Total=('Churn', 'count')
    ).reset_index()
    churn_by_age['Churn_Rate'] = churn_by_age['Churned'] / churn_by_age['Total']

    churn_by_year = product_df.groupby(product_df['Date'].apply(lambda x: x.year)).agg(
        Churned=('Churn', lambda x: (x == 1).sum()),
        Total=('Churn', 'count')
    ).reset_index()
    churn_by_year['Churn_Rate'] = churn_by_year['Churned'] / churn_by_year['Total']

    return_by_age = product_df.groupby('Age brackets').agg(
        Returns=('Returns', lambda x: (x == 1).sum()),
        Total=('Returns', 'count')
    ).reset_index()
    return_by_age['Return_Rate'] = return_by_age['Returns'] / return_by_age['Total']

    total_purchase_by_category = product_df.groupby('Category').agg(
        Total_Purchase_Amount=('Purch_Amt', 'sum')
    ).reset_index()

    grouped_df = product_df.groupby('Gender').agg(
        Churned=('Churn', lambda x: (x == 1).sum()),
        Retained=('Churn', lambda x: (x == 0).sum())
    ).reset_index()

    # Create a pivot table for churn and retention based on Category and year
    pivot_table = product_df.pivot_table(
        index=['Category', product_df['Date'].apply(lambda x: x.year)],
        values='Churn',
        aggfunc=['sum', 'count']
    ).reset_index()
    pivot_table.columns = ['Category', 'Year', 'Churned', 'Total']
    pivot_table['Retained'] = pivot_table['Total'] - pivot_table['Churned']

    # Save analysis results to the Excel file
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
        churn_by_age.to_excel(writer, sheet_name='Churn_by_Age', index=False)
        churn_by_year.to_excel(writer, sheet_name='Churn_by_Year', index=False)
        return_by_age.to_excel(writer, sheet_name='Return_by_Age', index=False)
        total_purchase_by_category.to_excel(writer, sheet_name='Purchase_by_Category', index=False)
        grouped_df.to_excel(writer, sheet_name='Gender_Churn_Retained', index=False)
        pivot_table.to_excel(writer, sheet_name='Churn_Retained_by_Category_Year', index=False)

    # Create Dashboard
    wb = load_workbook(file_name)
    if 'Dashboard' not in wb.sheetnames:
        wb.create_sheet('Dashboard')
    sheet = wb['Dashboard']

    # Add title
    sheet.merge_cells('A1:R4')
    cell = sheet.cell(row=1, column=1)
    cell.value = 'E-commerce Purchase Activity Dashboard'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(b=True, color="F8F8F8", size=46)
    cell.fill = PatternFill("solid", fgColor="2591DB")

    # Add Churn by Age Chart
    churn_age_sheet = wb['Churn_by_Age']
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Churn Rate by Age Bracket"
    chart1.y_axis.title = 'Churn Rate'
    chart1.x_axis.title = 'Age Bracket'
    data1 = Reference(churn_age_sheet, min_col=4, min_row=1, max_row=4, max_col=4)
    cats1 = Reference(churn_age_sheet, min_col=1, min_row=2, max_row=4)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    sheet.add_chart(chart1, "A5")

    # Add Churn by Year Chart
    churn_year_sheet = wb['Churn_by_Year']
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "Churn Rate by Year"
    chart2.y_axis.title = 'Churn Rate'
    chart2.x_axis.title = 'Year'
    data2 = Reference(churn_year_sheet, min_col=3, min_row=1, max_row=4, max_col=3)
    cats2 = Reference(churn_year_sheet, min_col=1, min_row=2, max_row=4)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    sheet.add_chart(chart2, "J5")

    # Add Return Rate by Age Chart
    return_age_sheet = wb['Return_by_Age']
    chart3 = BarChart()
    chart3.type = "col"
    chart3.style = 10
    chart3.title = "Return Rate by Age Bracket"
    chart3.y_axis.title = 'Return Rate'
    chart3.x_axis.title = 'Age Bracket'
    data3 = Reference(return_age_sheet, min_col=4, min_row=1, max_row=4, max_col=4)
    cats3 = Reference(return_age_sheet, min_col=1, min_row=2, max_row=4)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    sheet.add_chart(chart3, "A20")

    # Add Total Purchase by Category Chart
    purchase_category_sheet = wb['Purchase_by_Category']
    chart4 = BarChart()
    chart4.type = "col"
    chart4.style = 10
    chart4.title = "Total Purchase Amount by Category"
    chart4.y_axis.title = 'Total Purchase Amount'
    chart4.x_axis.title = 'Category'
    data4 = Reference(purchase_category_sheet, min_col=2, min_row=1, max_row=4, max_col=2)
    cats4 = Reference(purchase_category_sheet, min_col=1, min_row=2, max_row=4)
    chart4.add_data(data4, titles_from_data=True)
    chart4.set_categories(cats4)
    sheet.add_chart(chart4, "J20")

    # Add Line Chart for Churn and Retention by Category and Year
    churn_retained_sheet = wb['Churn_Retained_by_Category_Year']
    chart5 = LineChart()
    chart5.title = "Churn and Retention by Category and Year"
    chart5.y_axis.title = 'Count'
    chart5.x_axis.title = 'Year'

    data5 = Reference(churn_retained_sheet, min_col=3, min_row=1, max_row=pivot_table.shape[0]+1, max_col=5)
    cats5 = Reference(churn_retained_sheet, min_col=2, min_row=2, max_row=pivot_table.shape[0]+1)
    chart5.add_data(data5, titles_from_data=True)
    chart5.set_categories(cats5)
    sheet.add_chart(chart5, "A35")

    # Save the workbook
    wb.save(file_name)
else:
    print(f"File {file_name} does not exist.")